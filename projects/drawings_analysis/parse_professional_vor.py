#!/usr/bin/env python3
"""
Парсинг ведомости от сметчика и сравнение с ВОР
"""

import openpyxl
import json

# Читаем ведомость от сметчика
wb = openpyxl.load_workbook('/root/.openclaw/media/inbound/Ведомость_объемов_работ_v7---9666bab8-d943-46de-8aef-44c40d1d97f4.xlsx')
ws = wb.active

# Структурируем данные
sections = {}
current_section = ''
positions = []

for row in range(1, ws.max_row + 1):
    col_a = ws.cell(row=row, column=1).value  # № п/п
    col_b = ws.cell(row=row, column=2).value  # № в ЛСР
    col_c = ws.cell(row=row, column=3).value  # Наименование
    col_d = ws.cell(row=row, column=4).value  # Ед. изм.
    col_e = ws.cell(row=row, column=5).value  # Кол-во
    col_f = ws.cell(row=row, column=6).value  # Ссылки на чертежи
    col_g = ws.cell(row=row, column=7).value  # Формула расчёта
    
    # Пропускаем пустые строки
    if not any([col_a, col_b, col_c, col_d, col_e]):
        continue
    
    # Разделы
    if col_c and isinstance(col_c, str) and 'Раздел' in col_c:
        current_section = col_c
        sections[current_section] = []
    elif col_a and (isinstance(col_a, int) or (isinstance(col_a, str) and str(col_a).strip().isdigit())):
        # Позиция
        pos = {
            'npp': str(col_a).strip(),
            'lsr': str(col_b) if col_b else '',
            'name': str(col_c).strip() if col_c else '',
            'unit': str(col_d).strip() if col_d else '',
            'qty': float(col_e) if isinstance(col_e, (int, float)) else str(col_e),
            'drawing': str(col_f) if col_f else '',
            'formula': str(col_g) if col_g else ''
        }
        positions.append(pos)
        if current_section:
            sections[current_section].append(pos)

print(f"Всего позиций: {len(positions)}")
print(f"Разделов: {len(sections)}")
print()

# Выводим разделы и позиции
for section, items in sections.items():
    print(f"\n=== {section} ===")
    print(f"Позиций: {len(items)}")
    for item in items[:10]:  # Первые 10 позиций каждого раздела
        print(f"  {item['npp']}. {item['name'][:70]} | {item['unit']} | {item['qty']}")
    if len(items) > 10:
        print(f"  ... и ещё {len(items) - 10} позиций")

# Сохраняем в JSON
with open('/root/.openclaw/workspace/projects/drawings_analysis/ведомость_сметчика.json', 'w', encoding='utf-8') as f:
    json.dump({'sections': sections, 'all_positions': positions}, f, ensure_ascii=False, indent=2)

print(f"\n\nСохранено в JSON: /root/.openclaw/workspace/projects/drawings_analysis/ведомость_сметчика.json")
