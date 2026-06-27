import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ЛСР БИМ (Методика 2)"

# Стили
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
header_font = Font(color='FFFFFF', bold=True, size=9)
section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
red_font = Font(color='C00000', bold=True)

# ===== ШАПКА =====
ws['A1'] = 'ЛОКАЛЬНЫЙ СМЕТНЫЙ РАСЧЕТ (СМЕТА) №'
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:N1')

ws['A2'] = 'Наименование программного продукта:'
ws['B2'] = 'ГРАНД-Смета, версия 2026.1'
ws['A3'] = 'Наименование редакции сметных нормативов:'
ws['B3'] = 'Приказ Минстроя России от 26.12.2019 № 876/пр; от 04.08.2020 № 421/пр; от 21.12.2020 № 812/пр; от 11.12.2020 № 774/пр'
ws['A4'] = 'Реквизиты письма Минстроя об индексах:'
ws['B4'] = '[УКАЗАТЬ ПИСЬМО ДЛЯ РЕГИОНА И КВАРТАЛА]'
ws['B4'].font = red_font
ws['A5'] = 'Реквизиты НПА об оплате труда:'
ws['B5'] = '[УКАЗАТЬ ПРИКАЗ/РЕГИОНАЛЬНЫЙ ДОКУМЕНТ]'
ws['B5'].font = red_font
ws['A6'] = 'Субъект РФ:'
ws['B6'] = '38. Иркутская область'
ws['A7'] = 'Зона:'
ws['B7'] = '1'
ws['A8'] = 'Объект:'
ws['B8'] = '[НАИМЕНОВАНИЕ ОБЪЕКТА]'
ws['B8'].font = red_font

# ===== ШАПКА ТАБЛИЦЫ =====
row_header = 10
headers = [
    '№ п/п', 'Обоснование', 'Наименование работ и затрат',
    'Ед.изм.', 'Кол-во', 'Коэфф.', 'Кол-во с учетом коэфф.',
    'Сметная стоимость в базисном уровне цен, руб.',
    'Индексы', 'Сметная стоимость в текущем уровне цен, руб.'
]

# Подзаголовки
ws.cell(row=row_header, column=1, value='1')
ws.cell(row=row_header, column=2, value='2')
ws.cell(row=row_header, column=3, value='3')
ws.cell(row=row_header, column=4, value='4')
ws.cell(row=row_header, column=5, value='5')
ws.cell(row=row_header, column=6, value='6')
ws.cell(row=row_header, column=7, value='7')
ws.cell(row=row_header, column=8, value='8')
ws.cell(row=row_header, column=9, value='9')
ws.cell(row=row_header, column=10, value='10')

for col in range(1, 11):
    c = ws.cell(row=row_header, column=col)
    c.fill = header_fill
    c.font = header_font
    c.border = thin_border
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# ===== СТРУКТУРА ПОЗИЦИИ =====
# Пример позиции (вентилятор радиальный)
row = 12
ws.cell(row=row, column=1, value='1')
ws.cell(row=row, column=2, value='ФЕР20-03-001-03')
ws.cell(row=row, column=3, value='Установка вентиляторов радиальных массой: свыше 0,12 до 0,2 т')
ws.cell(row=row, column=4, value='шт')
ws.cell(row=row, column=5, value=1)
ws.cell(row=row, column=6, value=1)
ws.cell(row=row, column=7, value='=E12*F12')

# Строки ресурсов
row_res = row + 1
ws.cell(row=row_res, column=2, value='421/пр_2020_п.58_пп.б')
ws.cell(row=row_res, column=3, value='Для работ аналогичных новому строительству: ОЗП=1.15; ЭМ=1.25; ЗПМ=1.25; ТЗ=1.15; ТЗМ=1.25')
ws.cell(row=row_res, column=3).font = Font(italic=True, size=8)

row_res += 1
ws.cell(row=row_res, column=2, value='421/пр_2020_прил.10_т.5_п.1.2_гр.3')
ws.cell(row=row_res, column=3, value='Ремонт в действующем здании: ОЗП=1.35; ЭМ=1.35; ЗПМ=1.35; ТЗ=1.35; ТЗМ=1.35')
ws.cell(row=row_res, column=3).font = Font(italic=True, size=8)

# Строки статей затрат
row_res += 1
ws.cell(row=row_res, column=2, value='1')
ws.cell(row=row_res, column=3, value='ОТ (Оплата труда)')
ws.cell(row=row_res, column=8, value=79.84).number_format = '0.00'
ws.cell(row=row_res, column=9, value='[INDEX_ZP]')
ws.cell(row=row_res, column=9).font = red_font
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*I' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=2, value='2')
ws.cell(row=row_res, column=3, value='ЭМ (Эксплуатация машин)')
ws.cell(row=row_res, column=8, value=21.95).number_format = '0.00'
ws.cell(row=row_res, column=9, value='[INDEX_MACH]')
ws.cell(row=row_res, column=9).font = red_font
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*I' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=2, value='3')
ws.cell(row=row_res, column=3, value='в т.ч. ОТм (оплата труда машинистов)')
ws.cell(row=row_res, column=8, value=2.11).number_format = '0.00'
ws.cell(row=row_res, column=9, value='[INDEX_ZP]')
ws.cell(row=row_res, column=9).font = red_font
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*I' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=2, value='4')
ws.cell(row=row_res, column=3, value='М (Материалы)')
ws.cell(row=row_res, column=8, value=21.14).number_format = '0.00'
ws.cell(row=row_res, column=9, value='[INDEX_MAT]')
ws.cell(row=row_res, column=9).font = red_font
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*I' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=3, value='ЗТ (Затраты труда рабочих)')
ws.cell(row=row_res, column=4, value='чел.-ч')
ws.cell(row=row_res, column=5, value=9.36).number_format = '0.00'
ws.cell(row=row_res, column=6, value=1.55)
ws.cell(row=row_res, column=7, value='=E' + str(row_res) + '*F' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=3, value='ЗТм (Затраты труда машинистов)')
ws.cell(row=row_res, column=4, value='чел.-ч')
ws.cell(row=row_res, column=5, value=0.17).number_format = '0.00'
ws.cell(row=row_res, column=6, value=1.69)
ws.cell(row=row_res, column=7, value='=E' + str(row_res) + '*F' + str(row_res)).number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=3, value='Итого по расценке')
ws.cell(row=row_res, column=3).font = Font(bold=True)
ws.cell(row=row_res, column=8, value='=SUM(H' + str(row_res-6) + ':H' + str(row_res-3) + ')').number_format = '0.00'
ws.cell(row=row_res, column=10, value='=SUM(J' + str(row_res-6) + ':J' + str(row_res-3) + ')').number_format = '0.00'

row_res += 1
ws.cell(row=row_res, column=3, value='ФОТ (Фонд оплаты труда)')
ws.cell(row=row_res, column=3).font = Font(bold=True)
ws.cell(row=row_res, column=8, value='=H' + str(row_res-1)).number_format = '0.00'
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*[INDEX_ZP]').number_format = '0.00'
ws.cell(row=row_res, column=10).font = red_font

# НР и СП
row_res += 1
ws.cell(row=row_res, column=2, value='Пр/812-xxx')
ws.cell(row=row_res, column=3, value='НР [шифр работ]')
ws.cell(row=row_res, column=4, value='%')
ws.cell(row=row_res, column=5, value='[НР%]')
ws.cell(row=row_res, column=5).font = red_font
ws.cell(row=row_res, column=8, value='=H' + str(row_res-1) + '*E' + str(row_res) + '/100').number_format = '0.00'
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*[INDEX_ZP]').number_format = '0.00'
ws.cell(row=row_res, column=10).font = red_font

row_res += 1
ws.cell(row=row_res, column=2, value='Пр/774-xxx')
ws.cell(row=row_res, column=3, value='СП [шифр работ]')
ws.cell(row=row_res, column=4, value='%')
ws.cell(row=row_res, column=5, value='[СП%]')
ws.cell(row=row_res, column=5).font = red_font
ws.cell(row=row_res, column=8, value='=H' + str(row_res-2) + '*E' + str(row_res) + '/100').number_format = '0.00'
ws.cell(row=row_res, column=10, value='=H' + str(row_res) + '*[INDEX_ZP]').number_format = '0.00'
ws.cell(row=row_res, column=10).font = red_font

row_res += 1
ws.cell(row=row_res, column=3, value='Всего по позиции')
ws.cell(row=row_res, column=3).font = Font(bold=True, size=10)
ws.cell(row=row_res, column=8, value='=H' + str(row_res-3) + '+H' + str(row_res-2) + '+H' + str(row_res-1)).number_format = '0.00'
ws.cell(row=row_res, column=8).fill = section_fill
ws.cell(row=row_res, column=10, value='=J' + str(row_res-3) + '+J' + str(row_res-2) + '+J' + str(row_res-1)).number_format = '0.00'
ws.cell(row=row_res, column=10).fill = section_fill

# ===== ИТОГИ СМЕТЫ =====
row_total = row_res + 2
ws.cell(row=row_total, column=3, value='ИТОГО ПРЯМЫЕ ЗАТРАТЫ')
ws.cell(row=row_total, column=3).font = Font(bold=True, size=11)
ws.cell(row=row_total, column=3).fill = section_fill
ws.cell(row=row_total, column=8, value='[СУММА ПРЯМЫХ]').number_format = '0.00'
ws.cell(row=row_total, column=8).fill = section_fill
ws.cell(row=row_total, column=10, value='[СУММА В ТЕКУЩИХ]').number_format = '0.00'
ws.cell(row=row_total, column=10).fill = section_fill

row_total += 1
ws.cell(row=row_total, column=3, value='ИТОГО НР')
ws.cell(row=row_total, column=8, value='[СУММА НР]').number_format = '0.00'
ws.cell(row=row_total, column=10, value='[СУММА НР ТЕКУЩИХ]').number_format = '0.00'

row_total += 1
ws.cell(row=row_total, column=3, value='ИТОГО СП')
ws.cell(row=row_total, column=8, value='[СУММА СП]').number_format = '0.00'
ws.cell(row=row_total, column=10, value='[СУММА СП ТЕКУЩИХ]').number_format = '0.00'

row_total += 1
ws.cell(row=row_total, column=3, value='ВСЕГО ПО СМЕТЕ')
ws.cell(row=row_total, column=3).font = Font(bold=True, size=12)
ws.cell(row=row_total, column=3).fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
ws.cell(row=row_total, column=8, value='=H' + str(row_total-3) + '+H' + str(row_total-2) + '+H' + str(row_total-1)).number_format = '0.00'
ws.cell(row=row_total, column=8).fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
ws.cell(row=row_total, column=10, value='=J' + str(row_total-3) + '+J' + str(row_total-2) + '+J' + str(row_total-1)).number_format = '0.00'
ws.cell(row=row_total, column=10).fill = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')

# ===== ШИРИНЫ КОЛОНОК =====
ws.column_dimensions['A'].width = 6
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 60
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 10
ws.column_dimensions['G'].width = 12
ws.column_dimensions['H'].width = 14
ws.column_dimensions['I'].width = 12
ws.column_dimensions['J'].width = 14

# ===== ЛИСТ ПОЯСНЕНИЙ =====
ws_notes = wb.create_sheet("Пояснения")
ws_notes['A1'] = 'ПОЯСНЕНИЯ К ШАБЛОНУ БИМ'
ws_notes['A1'].font = Font(bold=True, size=14)

notes = [
    ('', ''),
    ('БИМ (Базисно-Индексный Метод)', ''),
    ('Базисные цены: 01.01.2000 (руб. 2000 г.)', ''),
    ('', ''),
    ('ПОПРАВКИ (две применяются последовательно):', ''),
    ('1. П.58 пр.б Приказа 421/пр', 'Для работ аналогичных новому строительству: ОЗП=1.15; ЭМ=1.25; ЗПМ=1.25; ТЗ=1.15; ТЗМ=1.25'),
    ('2. Прил.10 т.5 п.1.2 гр.3 Приказа 421/пр', 'Ремонт в действующем здании: ОЗП=1.35; ЭМ=1.35; ЗПМ=1.35; ТЗ=1.35; ТЗМ=1.35'),
    ('', ''),
    ('ИТОГОВЫЕ КОЭФФИЦИЕНТЫ:', ''),
    ('ОТ (оплата труда)', '1.15 × 1.35 = 1.5525 ≈ 1.55'),
    ('ЭМ (эксплуатация машин)', '1.25 × 1.35 = 1.6875 ≈ 1.69'),
    ('М (материалы)', '1.0 × 1.0 = 1.0 (или по индексу материалов)'),
    ('', ''),
    ('ИНДЕКСЫ (по статьям, из письма Минстроя):', ''),
    ('INDEX_ZP (ОТ)', 'Индекс оплаты труда (обычно самый высокий)'),
    ('INDEX_MACH (ЭМ)', 'Индекс эксплуатации машин'),
    ('INDEX_MAT (М)', 'Индекс материалов'),
    ('', ''),
    ('НР и СП:', ''),
    ('НР считается от ФОТ (фонда оплаты труда) с коэффициентом по шифру работ', ''),
    ('СП считается от ФОТ с коэффициентом по шифру работ', ''),
    ('Шифр для сантехнических работ (вентиляция): Пр/812-016.0-1 (НР=121%), Пр/774-016.0 (СП=72%)', ''),
    ('', ''),
    ('ФОТ = ОТ × К1.58 + ОТм × К1.58 + ... (все статьи оплаты труда)', ''),
    ('', ''),
    ('МАТЕРИАЛЫ (ФССЦ):', ''),
    ('Материалы добавляются ОТДЕЛЬНЫМИ позициями после работы', ''),
    ('Код ФССЦ: 19.x.x.x-xxxx', ''),
    ('Цена из ФССЦ умножается на индекс материалов', ''),
    ('', ''),
    ('ПОРЯДОК РАБОТЫ:', ''),
    ('1. Получить ВОР/ведомость дефектов', ''),
    ('2. Подобрать код ФЕР в ГРАНД-Смете или PDF сборнике', ''),
    ('3. Заполнить статьи затрат (ОТ, ЭМ, М, ЗТ, ЗТм)', ''),
    ('4. Применить поправки (п.58 + Прил.10)', ''),
    ('5. Добавить ФССЦ на материалы', ''),
    ('6. Рассчитать НР/СП от ФОТ по шифру', ''),
    ('7. Применить индексы по статьям', ''),
    ('8. Проверить eval_smeta.py', ''),
]

for i, (a, b) in enumerate(notes, 2):
    ws_notes.cell(row=i, column=1, value=a)
    ws_notes.cell(row=i, column=2, value=b)
    if a and a.endswith(':'):
        ws_notes.cell(row=i, column=1).font = Font(bold=True)

ws_notes.column_dimensions['A'].width = 50
ws_notes.column_dimensions['B'].width = 80

# ===== СОХРАНЕНИЕ =====
output = '/root/.openclaw/workspace/projects/smeta/ШАБЛОН_ЛСР_БИМ_Методика2.xlsx'
wb.save(output)
print(f'Сохранено: {output}')
