#!/usr/bin/env python3
"""
EVAL для проверки сметы против ВОР (Ведомости объёмов работ) или ведомости дефектов.

Правило: ВОР/ведомость дефектов — ЕДИНСТВЕННЫЙ источник объёмов работ.
Всё, что в ВОР → должно быть в смете.
Всё, что в смете → должно быть в ВОР.

Использование:
    python3 eval_smeta.py --smeta <файл_сметы.xlsx> --vor <файл_ВОР.xlsx>

Exit codes:
    0 — OK (WARN допустимы)
    1 — FAIL (критические ошибки)
"""

import argparse
import sys
import openpyxl
from dataclasses import dataclass
from typing import List, Optional


@dataclass
class VORItem:
    """Позиция из ВОР"""
    num: str           # № п/п
    name: str          # Наименование
    unit: str          # Ед. изм.
    qty: float         # Количество


@dataclass
class SmetaItem:
    """Позиция из сметы"""
    num: str           # №
    code: str          # Код ФЕР
    name: str          # Наименование
    unit: str          # Ед
    qty: float         # Кол-во
    status: str = ""   # "УТОЧНИТЬ" или пусто


def parse_vor(filepath: str) -> List[VORItem]:
    """Парсит ВОР из Excel. Ожидает колонки: №, Наименование, Ед, Кол-во"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    items = []
    
    # Ищем заголовок таблицы
    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and str(cell_a).strip().lower() in ['№ п/п', '№', 'n', 'номер']:
            header_row = row
            break
    
    if not header_row:
        print(f"FAIL: Не найден заголовок таблицы в ВОР {filepath}")
        sys.exit(1)
    
    for row in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        unit = ws.cell(row=row, column=3).value
        qty = ws.cell(row=row, column=4).value
        
        # Пропускаем пустые и заголовки разделов
        if not num or not name:
            continue
        if isinstance(num, str) and 'раздел' in num.lower():
            continue
        if isinstance(name, str) and len(name) > 100 and 'условия' in name.lower():
            continue  # Примечания/условия
            
        try:
            qty_val = float(qty) if qty else 0
        except (ValueError, TypeError):
            continue
            
        items.append(VORItem(
            num=str(num).strip(),
            name=str(name).strip(),
            unit=str(unit).strip() if unit else '',
            qty=qty_val
        ))
    
    return items


def parse_smeta(filepath: str) -> List[SmetaItem]:
    """Парсит смету из Excel. Ожидает колонки: №, Код ФЕР, Наименование, Ед, Кол-во"""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active
    items = []
    
    # Ищем заголовок таблицы
    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and str(cell_a).strip() == '№':
            header_row = row
            break
    
    if not header_row:
        print(f"FAIL: Не найден заголовок таблицы в смете {filepath}")
        sys.exit(1)
    
    for row in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value
        unit = ws.cell(row=row, column=4).value
        qty = ws.cell(row=row, column=5).value
        
        # Пропускаем итоги, пустые, заголовки
        if not num or not name:
            continue
        num_str = str(num).strip().lower()
        if any(x in num_str for x in ['итого', 'всего', 'накладные', 'прибыль', 'раздел']):
            continue
        if isinstance(name, str) and 'внимание' in name.lower():
            continue
            
        # Проверяем статус "УТОЧНИТЬ"
        status = ""
        if code and isinstance(code, str) and 'уточнить' in code.lower():
            status = "УТОЧНИТЬ"
        if name and isinstance(name, str) and 'уточнить' in name.lower():
            status = "УТОЧНИТЬ"
        
        try:
            if isinstance(qty, str) and qty.startswith('='):
                # Формула — не можем вычислить без data_only=True
                qty_val = 0
            else:
                qty_val = float(qty) if qty else 0
        except (ValueError, TypeError):
            qty_val = 0
            
        items.append(SmetaItem(
            num=str(num).strip(),
            code=str(code).strip() if code else '',
            name=str(name).strip(),
            unit=str(unit).strip() if unit else '',
            qty=qty_val,
            status=status
        ))
    
    return items


def check_vor_in_smeta(vor_items: List[VORItem], smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет, что всё из ВОР есть в смете"""
    errors = []
    smeta_names = [s.name.lower() for s in smeta_items]
    
    for vor in vor_items:
        found = False
        for smeta in smeta_items:
            # Проверяем похожесть названий
            vor_name = vor.name.lower()
            smeta_name = smeta.name.lower()
            
            # Простое совпадение по ключевым словам
            if any(word in smeta_name for word in vor_name.split()[:3]):
                found = True
                break
        
        if not found:
            errors.append(
                f"FAIL: Позиция из ВОР отсутствует в смете: "
                f"№{vor.num} '{vor.name}' ({vor.qty} {vor.unit})"
            )
    
    return errors


def check_smeta_in_vor(smeta_items: List[SmetaItem], vor_items: List[VORItem]) -> List[str]:
    """Проверяет, что в смете нет лишних позиций (не из ВОР)"""
    warnings = []
    vor_names = [v.name.lower() for v in vor_items]
    
    for smeta in smeta_items:
        if smeta.status == "УТОЧНИТЬ":
            continue  # Помеченные для уточнения — не ошибка
            
        found = False
        smeta_name = smeta.name.lower()
        
        for vor in vor_items:
            vor_name = vor.name.lower()
            if any(word in smeta_name for word in vor_name.split()[:3]):
                found = True
                break
        
        if not found:
            warnings.append(
                f"WARN: Позиция в смете не найдена в ВОР: "
                f"№{smeta.num} '{smeta.name}' ({smeta.qty} {smeta.unit}) — "
                f"возможно, лишняя или требует пометки 'УТОЧНИТЬ'"
            )
    
    return warnings


def check_quantities(vor_items: List[VORItem], smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет соответствие количеств"""
    errors = []
    
    for vor in vor_items:
        for smeta in smeta_items:
            vor_name = vor.name.lower()
            smeta_name = smeta.name.lower()
            
            if any(word in smeta_name for word in vor_name.split()[:3]):
                # Проверяем единицы измерения
                if vor.unit and smeta.unit:
                    vor_unit = vor.unit.lower().replace('.', '')
                    smeta_unit = smeta.unit.lower().replace('.', '')
                    
                    # Учитываем разные формы (шт/шт.)
                    if vor_unit != smeta_unit:
                        # Проверяем масштаб (кг vs 100кг)
                        if '100кг' in smeta_unit and 'кг' in vor_unit:
                            expected_qty = vor.qty / 100
                            if abs(smeta.qty - expected_qty) > 0.01:
                                errors.append(
                                    f"FAIL: Неверное количество в смете для '{smeta.name}': "
                                    f"ВОР={vor.qty} {vor.unit}, Смета={smeta.qty} {smeta.unit} "
                                    f"(ожидалось {expected_qty} {smeta.unit})"
                                )
                        else:
                            errors.append(
                                f"WARN: Разные единицы измерения для '{smeta.name}': "
                                f"ВОР={vor.unit}, Смета={smeta.unit}"
                            )
                
                # Проверяем количество (если единицы совпадают)
                elif abs(vor.qty - smeta.qty) > 0.01:
                    errors.append(
                        f"FAIL: Неверное количество в смете для '{smeta.name}': "
                        f"ВОР={vor.qty} {vor.unit}, Смета={smeta.qty} {smeta.unit}"
                    )
                break
    
    return errors


def check_utocheniya(smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет, что позиции с 'УТОЧНИТЬ' действительно требуют уточнения"""
    warnings = []
    
    for smeta in smeta_items:
        if smeta.status == "УТОЧНИТЬ":
            warnings.append(
                f"WARN: Позиция требует уточнения: №{smeta.num} '{smeta.name}' — "
                f"нужно проверить в ГРАНД-Смете"
            )
    
    return warnings


def check_nr_sp(smeta_items: List[SmetaItem], ws) -> List[str]:
    """Проверяет наличие НР и СП в итогах"""
    warnings = []
    
    has_nr = False
    has_sp = False
    
    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and isinstance(cell_a, str):
            if 'накладные' in cell_a.lower() or 'нр' in cell_a.lower():
                has_nr = True
            if 'прибыль' in cell_a.lower() or 'сп' in cell_a.lower():
                has_sp = True
    
    if not has_nr:
        warnings.append("WARN: В смете отсутствуют Накладные расходы (НР)")
    if not has_sp:
        warnings.append("WARN: В смете отсутствует Сметная прибыль (СП)")
    
    return warnings


def main():
    parser = argparse.ArgumentParser(description='Проверка сметы против ВОР')
    parser.add_argument('--smeta', required=True, help='Файл сметы (.xlsx)')
    parser.add_argument('--vor', required=True, help='Файл ВОР (.xlsx)')
    args = parser.parse_args()
    
    print(f"=== ПРОВЕРКА СМЕТЫ ===")
    print(f"Смета: {args.smeta}")
    print(f"ВОР:   {args.vor}")
    print()
    
    # Парсим файлы
    try:
        vor_items = parse_vor(args.vor)
        smeta_items = parse_smeta(args.smeta)
    except Exception as e:
        print(f"FAIL: Ошибка парсинга: {e}")
        sys.exit(1)
    
    print(f"Позиций в ВОР:   {len(vor_items)}")
    print(f"Позиций в смете: {len(smeta_items)}")
    print()
    
    # Проверки
    all_errors = []
    all_warnings = []
    
    # 1. Всё из ВОР есть в смете
    errors = check_vor_in_smeta(vor_items, smeta_items)
    all_errors.extend(errors)
    
    # 2. В смете нет лишнего
    warnings = check_smeta_in_vor(smeta_items, vor_items)
    all_warnings.extend(warnings)
    
    # 3. Количества совпадают
    errors = check_quantities(vor_items, smeta_items)
    all_errors.extend(errors)
    
    # 4. Позиции с УТОЧНИТЬ
    warnings = check_utocheniya(smeta_items)
    all_warnings.extend(warnings)
    
    # 5. НР и СП
    wb = openpyxl.load_workbook(args.smeta)
    ws = wb.active
    warnings = check_nr_sp(smeta_items, ws)
    all_warnings.extend(warnings)
    
    # Выводим результаты
    has_fail = False
    
    if all_errors:
        print("❌ ОШИБКИ (FAIL):")
        for e in all_errors:
            print(f"  {e}")
            if e.startswith("FAIL:"):
                has_fail = True
        print()
    
    if all_warnings:
        print("⚠️  ПРЕДУПРЕЖДЕНИЯ (WARN):")
        for w in all_warnings:
            print(f"  {w}")
        print()
    
    if not all_errors and not all_warnings:
        print("✅ ПРОВЕРКА ПРОЙДЕНА: Смета полностью соответствует ВОР")
        sys.exit(0)
    elif not has_fail:
        print("✅ ПРОВЕРКА ПРОЙДЕНА С ЗАМЕЧАНИЯМИ: WARN требуют внимания, но не критичны")
        sys.exit(0)
    else:
        print(f"❌ ПРОВЕРКА НЕ ПРОЙДЕНА: {len([e for e in all_errors if e.startswith('FAIL:')])} критических ошибок")
        sys.exit(1)


if __name__ == '__main__':
    main()
