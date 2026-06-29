#!/usr/bin/env python3
"""
Генератор шаблона ЛСР (локальная смета расходов) по БИМ-методике.

Создаёт Excel-файл с:
- Колонками по статьям: ОТ, ЭМ, М
- Формулами для текущих цен (индексы по статьям)
- НР/СП от ФОТ
- Итоговыми суммами

Использование:
    python3 create_bim_template.py --output смета_БИМ.xlsx --region "Иркутск" --year 2026 --quarter 2
"""

import argparse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_bim_template(output_path: str, region: str, year: int, quarter: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЛСР БИМ"

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Заголовок
    ws.merge_cells('A1:M1')
    ws['A1'] = f"ЛОКАЛЬНАЯ СМЕТА РАСХОДОВ (БИМ, Методика 2)"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:M2')
    ws['A2'] = f"Регион: {region} | {year} г., {quarter} квартал"
    ws['A2'].alignment = Alignment(horizontal='center')

    # Заголовки таблицы
    headers = [
        "№", "Код ФЕР", "Наименование", "Ед",
        "Кол-во",
        "ОТ (руб. 2000)", "ЭМ (руб. 2000)", "М (руб. 2000)",
        "Поправки К", "ОТ текущ.", "ЭМ текущ.", "М текущ.", "Всего текущ."
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

    # Примерные формулы для первой позиции (строка 5)
    # ОТ текущ. = ОТ баз. × К × INDEX_ZP
    # ЭМ текущ. = ЭМ баз. × К × INDEX_MACH
    # М текущ. = М баз. × К × INDEX_MAT
    # Всего = ОТ текущ. + ЭМ текущ. + М текущ.

    # Примечания с индексами
    ws.merge_cells('A3:M3')
    ws['A3'] = f"INDEX_ZP=62.7, INDEX_MACH=16.84, INDEX_MAT=8.42 | Поправки: п.58=1.15/1.25/1.25, Прил.10=1.35"
    ws['A3'].font = Font(italic=True, size=9)
    ws['A3'].alignment = Alignment(horizontal='center')

    # Итоговые строки (будут ниже, после позиций)
    # Строка 50 - примерно
    total_row = 50
    ws.cell(row=total_row, column=1, value="ИТОГО ПРЯМЫЕ ЗАТРАТЫ:")
    ws.cell(row=total_row, column=1).font = Font(bold=True)

    # НР и СП
    nr_row = total_row + 1
    ws.cell(row=nr_row, column=1, value="НР (от ФОТ, шифр Пр/812-016.0-1): 121%")
    ws.cell(row=nr_row, column=1).font = Font(bold=True, color="FF0000")

    sp_row = nr_row + 1
    ws.cell(row=sp_row, column=1, value="СП (от ФОТ, шифр Пр/774-016.0): 72%")
    ws.cell(row=sp_row, column=1).font = Font(bold=True, color="FF0000")

    grand_row = sp_row + 1
    ws.cell(row=grand_row, column=1, value="ВСЕГО ПО СМЕТЕ:")
    ws.cell(row=grand_row, column=1).font = Font(bold=True, size=12)

    # Ширины колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 14
    ws.column_dimensions['M'].width = 14

    # Примечание о методологии
    note_row = grand_row + 3
    ws.merge_cells(f'A{note_row}:M{note_row}')
    ws.cell(row=note_row, value="ПРИМЕЧАНИЕ: БИМ (базисно-индексный метод), Методика 2, приказ 421/пр. НР/СП считаются от ФОТ.")
    ws.cell(row=note_row).font = Font(italic=True, size=9)

    wb.save(output_path)
    print(f"Шаблон БИМ-сметы создан: {output_path}")


if __name__ == '__main__':
    parser = argparse.ArgumentParser()
    parser.add_argument('--output', default='/tmp/ШАБЛОН_ЛСР_БИМ.xlsx')
    parser.add_argument('--region', default='Иркутск')
    parser.add_argument('--year', type=int, default=2026)
    parser.add_argument('--quarter', type=int, default=2)
    args = parser.parse_args()

    create_bim_template(args.output, args.region, args.year, args.quarter)
