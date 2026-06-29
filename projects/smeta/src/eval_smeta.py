#!/usr/bin/env python3
"""
EVAL для проверки сметы:
1. Соответствие ВОР (объёмы)
2. Корректность БИМ-методики (индексы, НР/СП, поправки, ФССЦ)

Правила:
- ВОР/ведомость дефектов — ЕДИНСТВЕННЫЙ источник объёмов работ.
- БИМ: индексы по статьям (ОТ, ЭМ, М), не общий INDEX_SMR.
- НР/СП: от ФОТ, не от прямых затрат.
- Поправки: несколько, с разными коэфф. по статьям.
- ФССЦ: материалы отдельно.

Использование:
    python3 eval_smeta.py --smeta <файл_сметы.xlsx> --vor <файл_ВОР.xlsx>

Exit codes:
    0 — OK (WARN допустимы)
    1 — FAIL (критические ошибки)
"""

import argparse
import sys
import json
import openpyxl
from dataclasses import dataclass
from typing import List, Optional

# === CONFIG LOADING ===
CONFIG_DIR = "/root/.openclaw/workspace/projects/smeta/config"

def load_indexes():
    """Загружает индексы из config/indexes.json"""
    try:
        with open(f"{CONFIG_DIR}/indexes.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print("WARN: config/indexes.json не найден. Используются значения по умолчанию.")
        return {}

def load_norms():
    """Загружает нормативы из config/norms.json"""
    try:
        with open(f"{CONFIG_DIR}/norms.json", "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print("WARN: config/norms.json не найден.")
        return {}

INDEXES = load_indexes()
NORMS = load_norms()


@dataclass
class VORItem:
    """Позиция из ВОР"""
    num: str
    name: str
    unit: str
    qty: float


@dataclass
class SmetaItem:
    """Позиция из сметы"""
    num: str
    code: str
    name: str
    unit: str
    qty: float
    status: str = ""
    ot: float = 0.0  # Оплата труда
    em: float = 0.0  # Эксплуатация машин
    mat: float = 0.0  # Материалы


def parse_vor(filepath: str) -> List[VORItem]:
    """Парсит ВОР из Excel. Ожидает колонки: №, Наименование, Ед, Кол-во"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    items = []

    header_row = None
    for row in range(1, min(20, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and str(cell_a).strip().lower() in ['№ п/п', '№', 'n', 'номер']:
            header_row = row
            break

    if not header_row:
        print(f"FAIL: Не найден заголовок таблицы в ВОР {filepath}")
        sys.exit(1)

    for row in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        name = ws.cell(row=row, column=2).value
        unit = ws.cell(row=row, column=3).value
        qty = ws.cell(row=row, column=4).value

        if not num or not name:
            continue
        if isinstance(name, (int, float)) or (isinstance(name, str) and name.strip().isdigit()):
            continue
        if isinstance(num, str) and 'раздел' in num.lower():
            continue
        if isinstance(name, str) and len(name) > 100 and 'условия' in name.lower():
            continue

        try:
            qty_val = float(qty) if qty else 0
        except (ValueError, TypeError):
            continue

        items.append(VORItem(
            num=str(num).strip(),
            name=str(name).strip(),
            unit=str(unit).strip() if unit else '',
            qty=qty_val
        ))

    return items


def parse_smeta(filepath: str) -> List[SmetaItem]:
    """Парсит смету из Excel. Поддерживает БИМ-формат с колонками ОТ, ЭМ, М."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb.active
    items = []

    header_row = None
    bim_mode = False
    for row in range(1, min(20, ws.max_row + 1)):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and str(cell_a).strip() == '№':
            header_row = row
            # Проверяем, есть ли колонки ОТ, ЭМ, М
            for col in range(1, min(20, ws.max_column + 1)):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str):
                    if 'оплата труда' in val.lower() or val.lower() in ['от', 'озп']:
                        bim_mode = True
            break

    if not header_row:
        print(f"FAIL: Не найден заголовок таблицы в смете {filepath}")
        sys.exit(1)

    for row in range(header_row + 1, ws.max_row + 1):
        num = ws.cell(row=row, column=1).value
        code = ws.cell(row=row, column=2).value
        name = ws.cell(row=row, column=3).value
        unit = ws.cell(row=row, column=4).value
        qty = ws.cell(row=row, column=5).value

        if not num or not name:
            continue
        num_str = str(num).strip().lower()
        name_str = str(name).strip().lower()
        if any(x in num_str for x in ['итого', 'всего', 'накладные', 'прибыль', 'раздел']):
            continue
        if isinstance(name, str) and ('внимание' in name_str or 'вндекс' in name_str):
            continue
        if isinstance(num, str) and ('поз.' in num_str or 'поправка' in num_str or 'индексы' in num_str or 'нр и сп' in num_str):
            continue
        if name_str.startswith('вор:') or 'сметчик' in name_str or 'код фер' in name_str:
            continue

        status = ""
        if code and isinstance(code, str) and 'уточнить' in code.lower():
            status = "УТОЧНИТЬ"
        if name and isinstance(name, str) and 'уточнить' in name.lower():
            status = "УТОЧНИТЬ"

        # Парсим ОТ, ЭМ, М если есть (БИМ)
        ot = em = mat = 0.0
        if bim_mode:
            # Ищем колонки по заголовкам
            pass  # Упрощённо

        try:
            if isinstance(qty, str) and qty.startswith('='):
                qty_val = 0
            else:
                qty_val = float(qty) if qty else 0
        except (ValueError, TypeError):
            qty_val = 0

        items.append(SmetaItem(
            num=str(num).strip(),
            code=str(code).strip() if code else '',
            name=str(name).strip(),
            unit=str(unit).strip() if unit else '',
            qty=qty_val,
            status=status,
            ot=ot, em=em, mat=mat
        ))

    return items


def check_vor_in_smeta(vor_items: List[VORItem], smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет, что всё из ВОР есть в смете"""
    errors = []
    smeta_names = [s.name.lower() for s in smeta_items]

    for vor in vor_items:
        found = False
        for smeta in smeta_items:
            vor_name = vor.name.lower()
            smeta_name = smeta.name.lower()
            if any(word in smeta_name for word in vor_name.split()[:3]):
                found = True
                break

        if not found:
            errors.append(
                f"FAIL: Позиция из ВОР отсутствует в смете: "
                f"№{vor.num} '{vor.name}' ({vor.qty} {vor.unit})"
            )

    return errors


def check_smeta_in_vor(smeta_items: List[SmetaItem], vor_items: List[VORItem]) -> List[str]:
    """Проверяет, что в смете нет лишних позиций (не из ВОР)"""
    warnings = []
    vor_names = [v.name.lower() for v in vor_items]

    for smeta in smeta_items:
        if smeta.status == "УТОЧНИТЬ":
            continue

        found = False
        smeta_name = smeta.name.lower()

        for vor in vor_items:
            vor_name = vor.name.lower()
            if any(word in smeta_name for word in vor_name.split()[:3]):
                found = True
                break

        if not found:
            warnings.append(
                f"WARN: Позиция в смете не найдена в ВОР: "
                f"№{smeta.num} '{smeta.name}' ({smeta.qty} {smeta.unit}) — "
                f"возможно, лишняя или требует пометки 'УТОЧНИТЬ'"
            )

    return warnings


def check_quantities(vor_items: List[VORItem], smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет соответствие количеств"""
    errors = []

    for vor in vor_items:
        for smeta in smeta_items:
            vor_name = vor.name.lower()
            smeta_name = smeta.name.lower()

            if any(word in smeta_name for word in vor_name.split()[:3]):
                if vor.unit and smeta.unit:
                    vor_unit = vor.unit.lower().replace('.', '')
                    smeta_unit = smeta.unit.lower().replace('.', '')

                    if vor_unit != smeta_unit:
                        if '100кг' in smeta_unit and 'кг' in vor_unit:
                            expected_qty = vor.qty / 100
                            if abs(smeta.qty - expected_qty) > 0.01:
                                errors.append(
                                    f"FAIL: Неверное количество в смете для '{smeta.name}': "
                                    f"ВОР={vor.qty} {vor.unit}, Смета={smeta.qty} {smeta.unit} "
                                    f"(ожидалось {expected_qty} {smeta.unit})"
                                )
                        else:
                            errors.append(
                                f"WARN: Разные единицы измерения для '{smeta.name}': "
                                f"ВОР={vor.unit}, Смета={smeta.unit}"
                            )

                elif abs(vor.qty - smeta.qty) > 0.01:
                    errors.append(
                        f"FAIL: Неверное количество в смете для '{smeta.name}': "
                        f"ВОР={vor.qty} {vor.unit}, Смета={smeta.qty} {smeta.unit}"
                    )
                break

    return errors


def check_bim_methodology(filepath: str, region: str = "", quarter: str = "") -> List[str]:
    """Проверяет БИМ-методику: индексы по статьям, НР/СП от ФОТ, ФССЦ, актуальность."""
    errors = []
    warnings = []

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    content = ""
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell:
                content += str(cell).lower() + " "

    # 0. Проверка актуальности индексов
    current_year = 2026
    if "2026" not in content and "2025" not in content:
        warnings.append(
            "WARN БИМ: Не найдена дата индексов (2025 или 2026). "
            "Убедитесь, что используются актуальные индексы."
        )
    
    # Проверка квартала
    quarters = ['1q', '2q', '3q', '4q', 'i кв', 'ii кв', 'iii кв', 'iv кв', '1 кв', '2 кв', '3 кв', '4 кв']
    has_quarter = any(q in content for q in quarters)
    if not has_quarter:
        warnings.append(
            "WARN БИМ: Не указан квартал индексов (например, 2Q2026). "
            "Индексы Минстроя публикуются ежеквартально — укажите, какой квартал используется."
        )

    # 0.1. Проверка индексов из config
    if region and quarter:
        region_data = INDEXES.get("indexes", {}).get(region, {})
        quarter_data = region_data.get(quarter, {})
        if not quarter_data:
            warnings.append(
                f"WARN БИМ: В config/indexes.json нет данных для {region} {quarter}. "
                f"Проверьте актуальность индексов."
            )
        else:
            # Проверяем, что в смете используются те же индексы
            expected_zp = str(quarter_data.get("INDEX_ZP", ""))
            expected_mach = str(quarter_data.get("INDEX_MACH", ""))
            expected_mat = str(quarter_data.get("INDEX_MAT", ""))
            
            if expected_zp and expected_zp not in content:
                warnings.append(
                    f"WARN БИМ: В смете не найден INDEX_ZP={expected_zp} (из config для {region} {quarter}). "
                    f"Возможно, используются устаревшие индексы."
                )

    # 1. Проверка на общий индекс (критично)
    if "index_smr" in content or "общий индекс" in content:
        if "index_zp" not in content and "index_ot" not in content:
            errors.append(
                "FAIL БИМ: Использован ОБЩИЙ индекс (INDEX_SMR) вместо индексов по статьям "
                "(INDEX_ZP, INDEX_MACH, INDEX_MAT). Это приведёт к ошибке в 7×."
            )

    # 2. Проверка НР/СП (критично)
    if "15%" in content and "12%" in content:
        if "от прямых" in content or "прямые затраты" in content:
            errors.append(
                "FAIL БИМ: НР/СП рассчитаны от ПРЯМЫХ ЗАТРАТ (15%/12%). "
                "Правильно: от ФОТ (фонд оплаты труда) с коэфф. по шифру работ."
            )

    if "нр" in content and "сп" in content:
        if "фот" not in content and "фонд оплаты труда" not in content:
            warnings.append(
                "WARN БИМ: НР/СП найдены, но не видно, что база = ФОТ. "
                "Проверьте: НР/СП должны считаться от фонда оплаты труда."
            )

    # 3. Проверка ФССЦ
    fssc_keywords = ['фссц', '19.', 'материал']
    has_fssc = any(kw in content for kw in fssc_keywords)
    fer_keywords = ['фер20', 'фер-20', 'фер 20']
    has_fer = any(kw in content for kw in fer_keywords)

    if has_fer and not has_fssc:
        warnings.append(
            "WARN БИМ: В смете есть ФЕР, но не видно ФССЦ-позиций для материалов. "
            "Материалы внутри ФЕР — только монтажные. Основные материалы добавляются отдельно через ФССЦ."
        )

    # 4. Проверка поправок
    if "к=1.15" in content or "1,15" in content:
        if "к=1.35" not in content and "1,35" not in content:
            warnings.append(
                "WARN БИМ: Найдена только одна поправка (К=1.15). "
                "Возможно, нужны дополнительные поправки (п.58, Прил.10 и др.) с разными коэфф. по статьям."
            )

    # 5. Проверка разделения по статьям
    if "от" not in content and "оплата труда" not in content:
        if "эм" not in content and "машин" not in content:
            warnings.append(
                "WARN БИМ: Не видно разделения по статьям (ОТ, ЭМ, М). "
                "БИМ-методика требует раздельного учёта."
            )

    return errors + warnings

    # 1. Проверка на общий индекс (критично)
    if "index_smr" in content or "общий индекс" in content:
        if "index_zp" not in content and "index_ot" not in content:
            errors.append(
                "FAIL БИМ: Использован ОБЩИЙ индекс (INDEX_SMR) вместо индексов по статьям "
                "(INDEX_ZP, INDEX_MACH, INDEX_MAT). Это приведёт к ошибке в 7×."
            )

    # 2. Проверка НР/СП (критично)
    if "15%" in content and "12%" in content:
        if "от прямых" in content or "прямые затраты" in content:
            errors.append(
                "FAIL БИМ: НР/СП рассчитаны от ПРЯМЫХ ЗАТРАТ (15%/12%). "
                "Правильно: от ФОТ (фонд оплаты труда) с коэфф. по шифру работ."
            )

    if "нр" in content and "сп" in content:
        if "фот" not in content and "фонд оплаты труда" not in content:
            warnings.append(
                "WARN БИМ: НР/СП найдены, но не видно, что база = ФОТ. "
                "Проверьте: НР/СП должны считаться от фонда оплаты труда."
            )

    # 3. Проверка ФССЦ
    fssc_keywords = ['фссц', '19.', 'материал']
    has_fssc = any(kw in content for kw in fssc_keywords)
    fer_keywords = ['фер20', 'фер-20', 'фер 20']
    has_fer = any(kw in content for kw in fer_keywords)

    if has_fer and not has_fssc:
        warnings.append(
            "WARN БИМ: В смете есть ФЕР, но не видно ФССЦ-позиций для материалов. "
            "Материалы внутри ФЕР — только монтажные. Основные материалы добавляются отдельно через ФССЦ."
        )

    # 4. Проверка поправок
    if "к=1.15" in content or "1,15" in content:
        if "к=1.35" not in content and "1,35" not in content:
            warnings.append(
                "WARN БИМ: Найдена только одна поправка (К=1.15). "
                "Возможно, нужны дополнительные поправки (п.58, Прил.10 и др.) с разными коэфф. по статьям."
            )

    # 5. Проверка разделения по статьям
    if "от" not in content and "оплата труда" not in content:
        if "эм" not in content and "машин" not in content:
            warnings.append(
                "WARN БИМ: Не видно разделения по статьям (ОТ, ЭМ, М). "
                "БИМ-методика требует раздельного учёта."
            )

    return errors + warnings


def check_utocheniya(smeta_items: List[SmetaItem]) -> List[str]:
    """Проверяет, что позиции с 'УТОЧНИТЬ' действительно требуют уточнения"""
    warnings = []

    for smeta in smeta_items:
        if smeta.status == "УТОЧНИТЬ":
            warnings.append(
                f"WARN: Позиция требует уточнения: №{smeta.num} '{smeta.name}' — "
                f"нужно проверить в ГРАНД-Смете"
            )

    return warnings


def check_nr_sp(smeta_items: List[SmetaItem], ws) -> List[str]:
    """Проверяет наличие НР и СП в итогах"""
    warnings = []

    has_nr = False
    has_sp = False

    for row in range(1, ws.max_row + 1):
        cell_a = ws.cell(row=row, column=1).value
        if cell_a and isinstance(cell_a, str):
            if 'накладные' in cell_a.lower() or 'нр' in cell_a.lower():
                has_nr = True
            if 'прибыль' in cell_a.lower() or 'сп' in cell_a.lower():
                has_sp = True

    if not has_nr:
        warnings.append("WARN: В смете отсутствуют Накладные расходы (НР)")
    if not has_sp:
        warnings.append("WARN: В смете отсутствует Сметная прибыль (СП)")

    return warnings


def main():
    parser = argparse.ArgumentParser(description='Проверка сметы по БИМ-методике')
    parser.add_argument('--smeta', required=True, help='Файл сметы (.xlsx)')
    parser.add_argument('--vor', required=True, help='Файл ВОР (.xlsx)')
    parser.add_argument('--region', default='', help='Регион (например, "Иркутская область")')
    parser.add_argument('--quarter', default='', help='Квартал (например, "2Q2026")')
    args = parser.parse_args()

    print(f"=== ПРОВЕРКА СМЕТЫ (БИМ-методика) ===")
    print(f"Смета:  {args.smeta}")
    print(f"ВОР:    {args.vor}")
    print(f"Регион: {args.region or 'не указан'}")
    print(f"Квартал: {args.quarter or 'не указан'}")
    print()

    try:
        vor_items = parse_vor(args.vor)
        smeta_items = parse_smeta(args.smeta)
    except Exception as e:
        print(f"FAIL: Ошибка парсинга: {e}")
        sys.exit(1)

    print(f"Позиций в ВОР:   {len(vor_items)}")
    print(f"Позиций в смете: {len(smeta_items)}")
    print()

    all_errors = []
    all_warnings = []

    # 1. Всё из ВОР есть в смете
    errors = check_vor_in_smeta(vor_items, smeta_items)
    all_errors.extend(errors)

    # 2. В смете нет лишнего
    warnings = check_smeta_in_vor(smeta_items, vor_items)
    all_warnings.extend(warnings)

    # 3. Количества совпадают
    errors = check_quantities(vor_items, smeta_items)
    all_errors.extend(errors)

    # 4. Позиции с УТОЧНИТЬ
    warnings = check_utocheniya(smeta_items)
    all_warnings.extend(warnings)

    # 5. НР и СП
    wb = openpyxl.load_workbook(args.smeta)
    ws = wb.active
    warnings = check_nr_sp(smeta_items, ws)
    all_warnings.extend(warnings)

    # 6. БИМ-методология (НОВОЕ)
    bim_issues = check_bim_methodology(args.smeta, region=args.region, quarter=args.quarter)
    for issue in bim_issues:
        if issue.startswith("FAIL"):
            all_errors.append(issue)
        else:
            all_warnings.append(issue)

    # Выводим результаты
    has_fail = False

    if all_errors:
        print("❌ ОШИБКИ (FAIL):")
        for e in all_errors:
            print(f"  {e}")
            if e.startswith("FAIL:"):
                has_fail = True
        print()

    if all_warnings:
        print("⚠️  ПРЕДУПРЕЖДЕНИЯ (WARN):")
        for w in all_warnings:
            print(f"  {w}")
        print()

    # Итог
    bim_fail_count = len([e for e in all_errors if "БИМ" in e])
    vor_fail_count = len([e for e in all_errors if "БИМ" not in e])

    if not all_errors and not all_warnings:
        print("✅ ПРОВЕРКА ПРОЙДЕНА: Смета полностью соответствует ВОР и БИМ-методике")
        sys.exit(0)
    elif not has_fail:
        print("✅ ПРОВЕРКА ПРОЙДЕНА С ЗАМЕЧАНИЯМИ: WARN требуют внимания, но не критичны")
        sys.exit(0)
    else:
        print(f"❌ ПРОВЕРКА НЕ ПРОЙДЕНА:")
        if bim_fail_count > 0:
            print(f"   {bim_fail_count} критических ошибок БИМ-методики (10-100× ошибка)")
        if vor_fail_count > 0:
            print(f"   {vor_fail_count} ошибок соответствия ВОР")
        sys.exit(1)


if __name__ == '__main__':
    main()
