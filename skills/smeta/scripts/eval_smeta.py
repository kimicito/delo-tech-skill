import openpyxl
import json
import sys
import argparse
from pathlib import Path


class SmetaEvaluator:
    """Проверка сметы на соответствие профессиональному стандарту."""

    def __init__(self, example_path="data/smeta_structure_example.json"):
        self.example = json.load(open(example_path, 'r', encoding='utf-8'))
        self.checks = []
        self.failures = []
        self.warnings = []

    def log(self, status, message):
        self.checks.append((status, message))
        if status == "FAIL":
            self.failures.append(message)
        elif status == "WARN":
            self.warnings.append(message)
        print(f"  [{status}] {message}")

    def check_structure(self, smeta_items):
        """Проверка структуры разделов."""
        print("\n=== 1. СТРУКТУРА РАЗДЕЛОВ ===")

        sections = {}
        for item in smeta_items:
            sec = item.get('section', 'БЕЗ_РАЗДЕЛА')
            if sec not in sections:
                sections[sec] = []
            sections[sec].append(item)

        expected_sections = list(self.example['structure'].keys())
        actual_sections = list(sections.keys())

        for sec in expected_sections:
            if sec in actual_sections:
                self.log("PASS", f"Раздел '{sec}' присутствует ({len(sections[sec])} позиций)")
            else:
                self.log("FAIL", f"Раздел '{sec}' ОТСУТСТВУЕТ! Критично для полноты сметы")

        for sec in actual_sections:
            if sec not in expected_sections and sec != 'БЕЗ_РАЗДЕЛА':
                self.log("WARN", f"Лишний раздел '{sec}' ({len(sections[sec])} позиций)")

    def check_subsections(self, smeta_items):
        """Проверка подразделов (фундаменты)."""
        print("\n=== 2. ПОДРАЗДЕЛЫ ФУНДАМЕНТОВ ===")

        fund_items = [i for i in smeta_items if 'Фундамент' in i.get('section', '')]
        subsections = {}
        for item in fund_items:
            sub = item.get('subsection', 'БЕЗ_ПОДРАЗДЕЛА')
            if sub not in subsections:
                subsections[sub] = []
            subsections[sub].append(item)

        expected_subs = self.example['structure']['Раздел 2. Фундаменты']['subsections']

        for sub_name, sub_data in expected_subs.items():
            if sub_name in subsections:
                actual_count = len(subsections[sub_name])
                expected_count = sub_data['items_count']
                self.log("PASS", f"Подраздел '{sub_name}': {actual_count} позиций (ожидалось ~{expected_count})")
            else:
                self.log("FAIL", f"Подраздел '{sub_name}' ОТСУТСТВУЕТ! Критично для полноты сметы")

    def check_fer_codes(self, smeta_items):
        """Проверка использования правильных ФЕР-кодов."""
        print("\n=== 3. ФЕР-КОДЫ (ГЭСН) ===")

        expected_codes = self.example.get('fer_codes_used', {})
        actual_codes = {}

        for item in smeta_items:
            code = item.get('code', '')
            if code and (code.startswith('ГЭСН') or code.startswith('ФЕР')):
                base_code = '-'.join(code.split('-')[:4])
                actual_codes[base_code] = actual_codes.get(base_code, 0) + 1

        for code, desc in expected_codes.items():
            if code in actual_codes:
                self.log("PASS", f"Код {code}: {actual_codes[code]} позиций")
            else:
                self.log("FAIL", f"Код {code} НЕ ИСПОЛЬЗОВАН! ({desc[:50]})")

        # Проверяем лишние коды
        for code in actual_codes:
            if code not in expected_codes:
                self.log("WARN", f"Лишний код {code}: {actual_codes[code]} позиций (проверить правильность)")

    def check_materials(self, smeta_items):
        """Проверка материалов (ПСЦ)."""
        print("\n=== 4. МАТЕРИАЛЫ (ПСЦ) ===")

        expected_materials = self.example.get('material_prices', {})
        actual_materials = {}

        for item in smeta_items:
            code = item.get('code', '')
            if code == 'ПСЦ':
                name = item.get('name', '')[:30]
                actual_materials[name] = actual_materials.get(name, 0) + 1

        if not actual_materials:
            self.log("FAIL", "Материалы (ПСЦ) не найдены! В смете должны быть материалы отдельно")
            return

        for mat_name in expected_materials:
            found = any(mat_name in name for name in actual_materials.keys())
            if found:
                self.log("PASS", f"Материал '{mat_name}' найден")
            else:
                self.log("WARN", f"Материал '{mat_name}' не найден — проверить, нужен ли")

        self.log("INFO", f"Всего материалов (ПСЦ): {len(actual_materials)} позиций")

    def check_nr_sp(self, smeta_items, ws_rows):
        """Проверка НР и СП."""
        print("\n=== 5. НР / СП (НАКЛАДНЫЕ РАСХОДЫ И СМЕТНАЯ ПРИБЫЛЬ) ===")

        # Проверяем НР и СП в строках (не только в позициях)
        has_nr = False
        has_sp = False
        has_profit = False

        for row in ws_rows:
            c = row.get('c', '')
            if not c:
                continue
            c_lower = str(c).lower()
            if 'нр' in c_lower and ('земляные' in c_lower or 'бетон' in c_lower or 'металл' in c_lower or 'тепло' in c_lower or 'кирпич' in c_lower or 'реконструк' in c_lower):
                has_nr = True
            if 'сп' in c_lower and ('земляные' in c_lower or 'бетон' in c_lower or 'металл' in c_lower or 'тепло' in c_lower or 'кирпич' in c_lower or 'реконструк' in c_lower):
                has_sp = True
            if 'сметная прибыль' in c_lower or 'прибыль' in c_lower:
                has_profit = True

        # Также проверяем в позициях
        for item in smeta_items:
            name_lower = item.get('name', '').lower()
            if 'нр' in name_lower:
                has_nr = True
            if 'сп' in name_lower:
                has_sp = True
            if 'прибыль' in name_lower:
                has_profit = True

        if has_nr:
            self.log("PASS", "НР (накладные расходы) присутствуют")
        else:
            self.log("FAIL", "НР ОТСУТСТВУЮТ! Критично — смета неполная")

        if has_sp:
            self.log("PASS", "СП (сметная прибыль) присутствует")
        else:
            self.log("FAIL", "СП ОТСУТСТВУЕТ! Критично — смета неполная")

        if has_profit:
            self.log("PASS", "Сметная прибыль присутствует")
        else:
            self.log("WARN", "Сметная прибыль не найдена — проверить")

    def check_completeness(self, smeta_items):
        """Общая проверка полноты."""
        print("\n=== 6. ПОЛНОТА СМЕТЫ ===")

        total_items = len(smeta_items)
        self.log("INFO", f"Всего позиций в смете: {total_items}")

        expected_total = 134
        if total_items < expected_total * 0.5:
            self.log("FAIL", f"Слишком мало позиций: {total_items}. Ожидалось ~{expected_total}")
        elif total_items < expected_total * 0.8:
            self.log("WARN", f"Мало позиций: {total_items}. Возможно, пропущены материалы или элементы")
        else:
            self.log("PASS", f"Количество позиций: {total_items} (в норме)")

        # Проверяем наличие позиций без кода
        no_code = [i for i in smeta_items if not i.get('code')]
        if no_code:
            self.log("WARN", f"Позиций без кода ФЕР: {len(no_code)}. Должны быть заполнены или помечены 'УТОЧНИТЬ'")
        else:
            self.log("PASS", "Все позиции имеют код ФЕР")

    def check_no_utochit(self, smeta_items):
        """Проверка на наличие 'УТОЧНИТЬ'."""
        print("\n=== 7. ПРОВЕРКА 'УТОЧНИТЬ' ===")

        utochit = [i for i in smeta_items if 'уточнить' in i.get('name', '').lower() or 'уточнить' in i.get('code', '').lower()]

        if utochit:
            self.log("FAIL", f"Найдено {len(utochit)} позиций с 'УТОЧНИТЬ'! Исправить перед отправкой")
            for item in utochit[:5]:
                self.log("FAIL", f"  - {item.get('name', '')[:60]}")
        else:
            self.log("PASS", "Позиций 'УТОЧНИТЬ' нет")

    def check_transport(self, smeta_items):
        """Проверка перевозки."""
        print("\n=== 8. ПЕРЕВОЗКА ===")

        transport = [i for i in smeta_items if 'перевозк' in i.get('name', '').lower() or 'груз' in i.get('name', '').lower()]

        if transport:
            self.log("PASS", f"Перевозка найдена: {len(transport)} позиций")
        else:
            self.log("WARN", "Перевозка не найдена — проверить, нужна ли (земляные работы)")

    def evaluate(self, smeta_items, ws_rows):
        """Запуск всех проверок."""
        print("=" * 60)
        print("EVAL SMETA — Проверка Локальной Сметной Расчёта")
        print("=" * 60)
        print(f"Эталон: {self.example['meta']['object']}")
        print(f"Версия: {self.example['meta']['date']}")
        print()

        self.check_structure(smeta_items)
        self.check_subsections(smeta_items)
        self.check_fer_codes(smeta_items)
        self.check_materials(smeta_items)
        self.check_nr_sp(smeta_items, ws_rows)
        self.check_completeness(smeta_items)
        self.check_no_utochit(smeta_items)
        self.check_transport(smeta_items)

        # Результат
        print("\n" + "=" * 60)
        print("РЕЗУЛЬТАТ")
        print("=" * 60)

        total = len(self.checks)
        passed = len([c for c in self.checks if c[0] == "PASS"])
        failed = len(self.failures)
        warnings = len(self.warnings)

        print(f"  Проверок: {total}")
        print(f"  PASS: {passed}")
        print(f"  FAIL: {failed}")
        print(f"  WARN: {warnings}")

        if failed > 0:
            print(f"\n  ❌ FAIL: {failed} критических ошибок. Смета требует доработки!")
            print("  Критические ошибки:")
            for f in self.failures:
                print(f"    - {f}")
            return 1
        elif warnings > 3:
            print(f"\n  ⚠️ WARNING: {warnings} замечаний. Рекомендуется проверить перед отправкой.")
            return 2
        else:
            print(f"\n  ✅ PASS: Смета прошла проверку. Можно отправлять.")
            return 0


def load_smeta_from_excel(path):
    """Загрузка сметы из Excel."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    items = []
    ws_rows = []
    current_section = 'БЕЗ_РАЗДЕЛА'
    current_subsection = None

    for i in range(1, ws.max_row + 1):
        row = ws[i]
        a = row[0].value
        b = row[1].value
        c = row[2].value
        h = row[7].value
        i_val = row[8].value
        j = row[9].value
        k = row[10].value
        l = row[11].value

        ws_rows.append({'a': a, 'b': b, 'c': c, 'row': i})

        # Раздел
        if a and isinstance(a, str) and 'Раздел' in a:
            current_section = a
            continue

        # Подраздел
        if a and isinstance(a, str) and not a.startswith('=') and c is None:
            a_str = a.strip()
            if not a_str.replace('.','',1).isdigit() and not a_str.isdigit():
                current_subsection = a_str
                continue

        # Позиция: номер п/п (число)
        is_position = False
        a_str = str(a).strip() if a else ''
        if a_str and (a_str.isdigit() or (a_str.replace('.','',1).isdigit() and a_str.count('.') <= 1)):
            is_position = True

        if is_position and c and isinstance(c, str):
            items.append({
                'section': current_section,
                'subsection': current_subsection,
                'n': a_str,
                'code': str(b).strip() if b else '',
                'name': c.strip(),
                'unit': str(h).strip() if h else '',
                'qty': i_val,
                'coef': j,
                'total_qty': k,
                'unit_cost': l,
                'row': i
            })

    return items, ws_rows


def main():
    parser = argparse.ArgumentParser(description='Eval Smeta — проверка сметы')
    parser.add_argument('--smeta', required=True, help='Путь к смете.xlsx')
    parser.add_argument('--example', default='data/smeta_structure_example.json', help='Путь к эталонной структуре')
    args = parser.parse_args()

    evaluator = SmetaEvaluator(args.example)
    smeta_items, ws_rows = load_smeta_from_excel(args.smeta)

    exit_code = evaluator.evaluate(smeta_items, ws_rows)
    sys.exit(exit_code)


if __name__ == '__main__':
    main()
